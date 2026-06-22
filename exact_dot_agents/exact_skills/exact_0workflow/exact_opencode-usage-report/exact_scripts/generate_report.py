#!/usr/bin/env python3
"""
opencode 用量报表生成器

从 opencode workspace API 拉取月度用量数据并生成 Excel 报表。

单位说明:
  - totalCost: 微美分 (µ¢), $1 = 100,000,000 µ¢
  - 定价表为 Go 计划执行价，含输入、输出、缓存读取三档
  - token 估算使用有效输入价 = cache_hit_rate × p_cache + (1 − cache_hit_rate) × p_in
  - 默认缓存命中率 98.6%（实测 100 条 Flash 会话），input:output 比 142:1

用法:
  1. 从浏览器导出 opencode.ai 的 cookie (Playwright JSON 格式)
  2. python3 scripts/generate_report.py --cookie-file cookies.json --output report.xlsx

依赖: httpx, openpyxl (uv venv && uv pip install httpx openpyxl)
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import httpx
except ImportError:
    print("缺少 httpx: uv pip install httpx openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl: uv pip install openpyxl")
    sys.exit(1)

# ── 定价表 (每 1M tokens, USD) ──
PRICING = {
    "deepseek-v4-flash":  {"in": 0.14,  "out": 0.28,  "cache": 0.0028},
    "deepseek-v4-pro":    {"in": 1.74,  "out": 3.48,  "cache": 0.0145},
    "mimo-v2.5":          {"in": 0.14,  "out": 0.28,  "cache": 0.0028},
    "mimo-v2.5-pro":      {"in": 1.74,  "out": 3.48,  "cache": 0.0145},
    "kimi-k2.6":          {"in": 0.95,  "out": 4.00,  "cache": 0.16},
    "kimi-k2.7-code":     {"in": 0.95,  "out": 4.00,  "cache": 0.19},
    "minimax-m3":         {"in": 0.30,  "out": 1.20,  "cache": 0.06},
    "qwen3.7-plus":       {"in": 0.40,  "out": 1.60,  "cache": 0.04},
    "qwen3.6-plus":       {"in": 0.50,  "out": 3.00,  "cache": 0.05},
    "glm-5.1":            {"in": 1.40,  "out": 4.40,  "cache": 0.26},
    "glm-5.2":            {"in": 1.40,  "out": 4.40,  "cache": 0.26},
    "glm-5":              {"in": 1.00,  "out": 3.20,  "cache": 0.20},
    "minimax-m2.7":       {"in": 0.30,  "out": 1.20,  "cache": 0.06},
    "minimax-m2.5":       {"in": 0.30,  "out": 1.20,  "cache": 0.06},
}

DISPLAY_NAME = {
    "deepseek-v4-flash": "DeepSeek V4 Flash",
    "deepseek-v4-pro": "DeepSeek V4 Pro",
    "mimo-v2.5": "MiMo V2.5",
    "mimo-v2.5-pro": "MiMo V2.5 Pro",
    "kimi-k2.6": "Kimi K2.6",
    "kimi-k2.7-code": "Kimi K2.7 Code",
    "minimax-m3": "MiniMax M3",
    "qwen3.7-plus": "Qwen3.7 Plus",
    "qwen3.6-plus": "Qwen3.6 Plus",
    "glm-5.2": "GLM-5.2",
    "glm-5.1": "GLM-5.1",
    "glm-5": "GLM-5",
}


# ── API ──
def build_rpc_body(workspace_id: str, year: int, month: int, tz: str = "+08:00"):
    return {
        "t": {"t": 9, "i": 0, "l": 4, "a": [
            {"t": 1, "s": workspace_id},
            {"t": 0, "s": year},
            {"t": 0, "s": month},
            {"t": 1, "s": tz},
        ], "o": 0},
        "f": 31, "m": [],
    }


def parse_rpc_response(text: str):
    records = []
    pattern = r'(\{date:"[^"]+",model:"[^"]+",totalCost:\d+,keyId:"[^"]+",plan:"[^"]+"\})'
    for match in re.finditer(pattern, text):
        entry = match.group()
        date = re.search(r'date:"([^"]+)"', entry)
        model = re.search(r'model:"([^"]+)"', entry)
        cost = re.search(r'totalCost:(\d+)', entry)
        plan = re.search(r'plan:"([^"]+)"', entry)
        if date and model and cost:
            records.append({
                "date": date.group(1),
                "model": model.group(1),
                "totalCost": int(cost.group(1)),
                "plan": plan.group(1) if plan else "",
            })
    return records


def fetch_usage(cookie_str: str, workspace_id: str, year: int, month: int):
    resp = httpx.post(
        "https://opencode.ai/_server",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Content-Type": "application/json",
            "Cookie": cookie_str,
            "x-server-id": "15702f3a12ff8bff357f8c2aa154a17e65b746d5f6b96adc9002c86ee0c15205",
            "x-server-instance": "server-fn:0",
        },
        json=build_rpc_body(workspace_id, year, month),
        timeout=30,
    )
    resp.raise_for_status()
    records = parse_rpc_response(resp.text)

    daily = {}
    for r in records:
        key = (r["date"], r["model"])
        daily[key] = daily.get(key, 0) + r["totalCost"]

    model_totals = {}
    for (d, m), cost in daily.items():
        model_totals[m] = model_totals.get(m, 0) + cost

    date_totals = {}
    for (d, m), cost in daily.items():
        date_totals[d] = date_totals.get(d, 0) + cost

    return {"daily": daily, "model_totals": model_totals, "date_totals": date_totals, "records": records}


# ── Token 估算 ──
def estimate_tokens(total_cost_usd: float, p_in: float, p_out: float, p_cache: float = 0.0,
                    ratio: float = 100.0, cache_hit_rate: float = 0.986):
    if ratio <= 0:
        return {"input": 0, "output": 0, "total": 0}
    eff_in = cache_hit_rate * p_cache + (1 - cache_hit_rate) * p_in
    out_tokens = total_cost_usd * 1_000_000 / (ratio * eff_in + p_out)
    in_tokens = ratio * out_tokens
    return {"input": int(in_tokens), "output": int(out_tokens), "total": int(in_tokens + out_tokens)}


# ── Excel ──
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER


def style_data(ws, r1, r2, ncol):
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if (r - r1) % 2 == 1:
                cell.fill = ALT_FILL


def generate_excel(data: dict, output: str = "opencode_usage_report.xlsx"):
    wb = Workbook()
    mt = data["model_totals"]
    dt = data["date_totals"]
    dl = data["daily"]

    sorted_models = sorted(mt.items(), key=lambda x: x[1], reverse=True)
    sorted_dates = sorted(dt.keys())
    total_spend = sum(mt.values())

    # ── Sheet 1: 费用总览 ──
    ws1 = wb.active
    ws1.title = "费用总览"
    ws1.sheet_properties.tabColor = "2F5496"

    ws1.cell(row=1, column=1, value="按模型汇总").font = Font(bold=True, size=14, color="2F5496")
    ws1.merge_cells("A1:G1")
    h1 = ["模型", "消费 (µ¢)", "消费 ($)", "占比", "估算输入 (M)", "估算输出 (M)", "估算总 Tokens (M)"]
    for i, h in enumerate(h1, 1):
        ws1.cell(row=2, column=i, value=h)
    style_header(ws1, 2, len(h1))

    row = 3
    for model, cost_ud in sorted_models:
        cost_usd = cost_ud / 100_000_000
        pct = cost_ud / total_spend * 100
        p = PRICING.get(model, {"in": 1.0, "out": 2.0, "cache": 0.0})
        ratio = 142 if model == "deepseek-v4-flash" else 100
        tok = estimate_tokens(cost_usd, p["in"], p["out"], p["cache"], ratio)
        ws1.cell(row=row, column=1, value=DISPLAY_NAME.get(model, model))
        ws1.cell(row=row, column=2, value=cost_ud).number_format = "#,##0"
        ws1.cell(row=row, column=3, value=round(cost_usd, 2)).number_format = "$#,##0.00"
        ws1.cell(row=row, column=4, value=round(pct, 1))
        ws1.cell(row=row, column=5, value=round(tok["input"] / 1e6, 2))
        ws1.cell(row=row, column=6, value=round(tok["output"] / 1e6, 2))
        ws1.cell(row=row, column=7, value=round(tok["total"] / 1e6, 2))
        row += 1

    total_tok = {"input": 0, "output": 0, "total": 0}
    for model, cost_ud in sorted_models:
        p = PRICING.get(model, {"in": 1.0, "out": 2.0, "cache": 0.0})
        ratio = 142 if model == "deepseek-v4-flash" else 100
        tok = estimate_tokens(cost_ud / 100_000_000, p["in"], p["out"], p["cache"], ratio)
        for k in total_tok:
            total_tok[k] += tok[k]

    for c in range(1, len(h1) + 1):
        ws1.cell(row=row, column=c).font = Font(bold=True, size=11)
        ws1.cell(row=row, column=c).fill = TOTAL_FILL
        ws1.cell(row=row, column=c).border = BORDER
    ws1.cell(row=row, column=1, value="合计")
    ws1.cell(row=row, column=2, value=total_spend).number_format = "#,##0"
    ws1.cell(row=row, column=3, value=round(total_spend / 100_000_000, 2)).number_format = "$#,##0.00"
    ws1.cell(row=row, column=4, value=100.0)
    ws1.cell(row=row, column=5, value=round(total_tok["input"] / 1e6, 2))
    ws1.cell(row=row, column=6, value=round(total_tok["output"] / 1e6, 2))
    ws1.cell(row=row, column=7, value=round(total_tok["total"] / 1e6, 2))
    style_data(ws1, 3, row, len(h1))

    pie = PieChart()
    pie.title = "模型消费占比 (USD)"
    pie.style = 10
    pie.add_data(Reference(ws1, min_col=3, min_row=2, max_row=row - 1), titles_from_data=True)
    pie.set_categories(Reference(ws1, min_col=1, min_row=3, max_row=row - 1))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.width, pie.height = 22, 14
    ws1.add_chart(pie, f"A{row + 3}")

    r2 = row + 20
    ws1.cell(row=r2, column=1, value="每日费用明细").font = Font(bold=True, size=14, color="2F5496")
    ws1.merge_cells(f"A{r2}:B{r2}")
    r2 += 1
    for i, h in enumerate(["日期", "费用 ($)"], 1):
        ws1.cell(row=r2, column=i, value=h)
    style_header(ws1, r2, 2)
    r2 += 1
    r_start = r2
    for dt_ in sorted_dates:
        ws1.cell(row=r2, column=1, value=dt_)
        ws1.cell(row=r2, column=2, value=round(dt[dt_] / 100_000_000, 2)).number_format = "$#,##0.00"
        r2 += 1
    style_data(ws1, r_start, r2 - 1, 2)

    bar = BarChart()
    bar.title, bar.style = "每日费用 (USD)", 10
    bar.y_axis.title = "USD"
    bar.add_data(Reference(ws1, min_col=2, min_row=r_start - 1, max_row=r2 - 1), titles_from_data=True)
    bar.set_categories(Reference(ws1, min_col=1, min_row=r_start, max_row=r2 - 1))
    bar.width, bar.height = 24, 14
    ws1.add_chart(bar, f"D{r_start - 2}")

    for c, w in zip("ABCDEFG", [22, 14, 14, 10, 16, 16, 18]):
        ws1.column_dimensions[c].width = w

    # ── Sheet 2: 每日明细 ──
    ws2 = wb.create_sheet("每日明细")
    ws2.sheet_properties.tabColor = "548235"
    all_models = [m for m, _ in sorted_models]
    h3 = ["日期"] + [DISPLAY_NAME.get(m, m) for m in all_models] + ["合计 ($)"]
    for i, h in enumerate(h3, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, len(h3))

    for ri, dt_ in enumerate(sorted_dates, 2):
        ws2.cell(row=ri, column=1, value=dt_)
        row_total = 0
        for ci, model in enumerate(all_models, 2):
            val = dl.get((dt_, model), 0) / 100_000_000
            ws2.cell(row=ri, column=ci, value=round(val, 4)).number_format = "$#,##0.0000"
            row_total += val
        ws2.cell(row=ri, column=len(h3), value=round(row_total, 4)).number_format = "$#,##0.0000"
    style_data(ws2, 2, len(sorted_dates) + 1, len(h3))

    bar2 = BarChart()
    bar2.title, bar2.style, bar2.y_axis.title, bar2.grouping = "每日各模型费用分组柱状图 (USD)", 10, "USD", "clustered"
    bar2.add_data(Reference(ws2, min_col=2, min_row=1, max_row=len(sorted_dates) + 1, max_col=len(all_models) + 1),
                  titles_from_data=True)
    bar2.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=len(sorted_dates) + 1))
    bar2.width, bar2.height = 28, 16
    ws2.add_chart(bar2, f"A{len(sorted_dates) + 4}")

    ws2.column_dimensions["A"].width = 14
    for ci in range(2, len(h3) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    # ── Sheet 3: 原始数据 ──
    ws3 = wb.create_sheet("原始数据")
    ws3.sheet_properties.tabColor = "BF8F00"
    records = data["records"]
    h4 = ["日期", "模型", "消费 (µ¢)", "消费 ($)", "Plan", "估算输入", "估算输出"]
    for i, h in enumerate(h4, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, 1, len(h4))

    for ri, r in enumerate(records, 2):
        cost_usd = r["totalCost"] / 100_000_000
        p = PRICING.get(r["model"], {"in": 1.0, "out": 2.0, "cache": 0.0})
        ratio = 142 if r["model"] == "deepseek-v4-flash" else 100
        tok = estimate_tokens(cost_usd, p["in"], p["out"], p["cache"], ratio)
        ws3.cell(row=ri, column=1, value=r["date"])
        ws3.cell(row=ri, column=2, value=r["model"])
        ws3.cell(row=ri, column=3, value=r["totalCost"]).number_format = "#,##0"
        ws3.cell(row=ri, column=4, value=round(cost_usd, 6)).number_format = "$#,##0.000000"
        ws3.cell(row=ri, column=5, value=r["plan"])
        ws3.cell(row=ri, column=6, value=tok["input"])
        ws3.cell(row=ri, column=7, value=tok["output"])
    style_data(ws3, 2, len(records) + 1, len(h4))
    for ci, w in enumerate([14, 24, 14, 14, 10, 16, 16], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output)
    return output


# ── CLI ──
def parse_cookie_file(path: str) -> str:
    text = Path(path).read_text()
    if text.strip().startswith("["):
        cookies = json.loads(text)
        parts = [f"{c['name']}={c['value']}" for c in cookies
                 if c.get("domain") in ("opencode.ai", ".opencode.ai")]
        return "; ".join(parts)
    return text.strip()


def main():
    ap = argparse.ArgumentParser(description="opencode 用量报表生成器")
    ap.add_argument("--cookie", help="Cookie 字符串")
    ap.add_argument("--cookie-file", help="Cookie JSON 文件路径")
    ap.add_argument("--workspace", default="wrk_01KST6NRHTMG20NM28XBJTGYFJ")
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--month", type=int, default=6)
    ap.add_argument("--output", default="opencode_usage_report.xlsx")
    args = ap.parse_args()

    if args.cookie:
        cookie = args.cookie
    elif args.cookie_file:
        cookie = parse_cookie_file(args.cookie_file)
    else:
        print("请指定 --cookie 或 --cookie-file")
        print("浏览器导出: await page.context().cookies() → 存为 JSON 传入 --cookie-file")
        sys.exit(1)

    print(f"获取 {args.year}年{args.month}月 用量数据...")
    data = fetch_usage(cookie, args.workspace, args.year, args.month)
    print(f"共 {len(data['records'])} 条记录, {len(data['model_totals'])} 个模型, {len(data['date_totals'])} 天")

    out = generate_excel(data, args.output)
    print(f"已保存: {out}")


if __name__ == "__main__":
    main()
